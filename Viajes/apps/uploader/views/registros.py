"""
Vistas relacionadas con la gestión de registros de pasajeros
"""
from apps.cuentas.roles import flujo_principal_required
from django.shortcuts import render, redirect
from django.contrib import messages
from django.core.paginator import Paginator
from django.db import models
from django.db.models import Count
from django.http import JsonResponse
from django.urls import reverse
from datetime import datetime, timedelta, time, timezone as dt_timezone
import logging
from decouple import config

from ..models import Registro, UploadBatch, Notificacion, CasoEspecial, TiemposAtencion
from ..utils.paises import get_paises

MESES_ES = {
    1: 'ene', 2: 'feb', 3: 'mar', 4: 'abr', 5: 'may', 6: 'jun',
    7: 'jul', 8: 'ago', 9: 'sep', 10: 'oct', 11: 'nov', 12: 'dic'
}

# Mapeo aeropuerto de salida (columna 起飞机场 del Excel) -> (ciudad, país)
# usado para autocompletar el PIN sin tener que tocar código cuando llega
# un vuelo nuevo. Para sumar una ruta agregar una entrada aquí.
ORIGEN_POR_AEROPUERTO = {
    'PEK': ('Pekín', 'China'),
    '北京': ('Pekín', 'China'),
}


def _detectar_vuelo_y_origen(registros):
    """Lee el primer arribo del queryset y devuelve (vuelo, ciudad, país).

    El número de vuelo se toma de UploadBatch (que ya lo extrae del Excel
    al subir el archivo) con fallback al campo del propio registro.
    Si el aeropuerto de salida no está en ORIGEN_POR_AEROPUERTO, se usa
    el código del aeropuerto como ciudad/país de fallback para no romper
    el texto del PIN.
    """
    primer = registros.select_related('batch').first()
    if primer is None:
        return '—', '—', '—'

    batch_vuelo = primer.batch.vuelo_numero if primer.batch_id else None
    vuelo = (batch_vuelo or primer.vuelo_numero or '').strip() or '—'

    aeropuerto = (primer.aeropuerto_salida or '').upper()
    for clave, (ciudad, pais) in ORIGEN_POR_AEROPUERTO.items():
        if clave.upper() in aeropuerto:
            return vuelo, ciudad, pais

    fallback = (primer.aeropuerto_salida or '—').strip() or '—'
    return vuelo, fallback, fallback
@flujo_principal_required
def update_registro(request, registro_id):
    """Vista para actualizar campos (TODOS pueden editar TODO)"""
    from urllib.parse import urlencode
    from django.urls import reverse
    
    if request.method == 'POST':
        try:
            registro = Registro.objects.get(id=registro_id)
            
            # ✅ SIN VALIDACIÓN DE PERMISOS - Todos pueden editar todo
            
            # Actualizar campos según la nueva lógica:
            # SR (Segunda Revisión) = segunda_revision
            # R (Rechazo) = rechazado
            # I (Internación) = internacion
            
            if 'segunda_revision' in request.POST:
                # Toggle Segunda Revisión
                registro.segunda_revision = request.POST.get('segunda_revision') == 'true'
                
                if not registro.segunda_revision:
                    # Si se desactiva SR, también desactivar R e I
                    registro.internacion = False
                    registro.rechazado = False
                    
            # R solo se puede activar si SR está activo
            elif 'rechazado' in request.POST:
                nuevo_valor = request.POST.get('rechazado') == 'true'

                # Si está intentando ACTIVAR R, validar que SR esté activo
                if nuevo_valor and not registro.segunda_revision:
                    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                        return JsonResponse({'success': False, 'error': '⚠️ Debes marcar SR antes de poder rechazar.'})
                    messages.warning(request, '⚠️ Debes marcar "Segunda Revisión (SR)" antes de poder rechazar.')
                    params = request.GET.copy()
                    params['highlight'] = str(registro_id)
                    redirect_url = reverse('admin_list') + '?' + urlencode(params)
                    return redirect(redirect_url)

                # Si la validación pasa (o está desactivando), actualizar
                registro.rechazado = nuevo_valor
                # Si se marca como Rechazo, desmarcar Internación
                if registro.rechazado:
                    registro.internacion = False


            # I solo se puede activar si SR está activo
            elif 'internacion' in request.POST:
                nuevo_valor = request.POST.get('internacion') == 'true'

                # Si está intentando ACTIVAR I, validar que SR esté activo
                if nuevo_valor and not registro.segunda_revision:
                    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                        return JsonResponse({'success': False, 'error': '⚠️ Debes marcar SR antes de marcar Internación.'})
                    messages.warning(request, '⚠️ Debes marcar "Segunda Revisión (SR)" antes de marcar Internación (I).')
                    params = request.GET.copy()
                    params['highlight'] = str(registro_id)
                    redirect_url = reverse('admin_list') + '?' + urlencode(params)
                    return redirect(redirect_url)

                # Si la validación pasa (o está desactivando), actualizar
                registro.internacion = nuevo_valor
                # Si se marca I, desmarcar Rechazo
                if registro.internacion:
                    registro.rechazado = False
            
            elif 'comentario' in request.POST:
                registro.comentario = request.POST.get('comentario')
            
            registro.save()

            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True,
                    'segunda_revision': registro.segunda_revision,
                    'rechazado': registro.rechazado,
                    'internacion': registro.internacion,
                })

            messages.success(request, ' Registro actualizado exitosamente.')

            # Mantener TODOS los parámetros GET que venían en la URL
            params = request.GET.copy()

            # Agregar el parámetro highlight
            params['highlight'] = str(registro_id)

            # Construir la URL completa manteniendo búsqueda, filtros, paginación, etc.
            redirect_url = reverse('admin_list') + '?' + urlencode(params)
            return redirect(redirect_url)
            
        except Registro.DoesNotExist:
            messages.error(request, ' Registro no encontrado.')
        except Exception as e:
            messages.error(request, f' Error al actualizar: {str(e)}')
    
    return redirect('admin_list')


@flujo_principal_required
def admin_list(request):
    """Vista para ver y modificar registros (TODOS VEN TODO)"""
    # TODOS ven TODOS los registros
    registros = Registro.objects.select_related('batch', 'batch__usuario').all()
    batches = UploadBatch.objects.all().order_by('-fecha_carga')
    
    # Filtro de búsqueda por documento o pasajero
    search = request.GET.get('search')
    if search:
        registros = registros.filter(
            models.Q(numero_documento__icontains=search) |
            models.Q(nombre_pasajero__icontains=search)
        )
    
    # Filtro por batch
    batch_id = request.GET.get('batch')
    if batch_id:
        registros = registros.filter(batch_id=batch_id)
    
    # Filtro por Segunda Revisión (SR)
    segunda_revision = request.GET.get('segunda_revision')
    if segunda_revision == 'true':
        registros = registros.filter(segunda_revision=True)
    
    # Filtro por Rechazo (R)
    rechazado = request.GET.get('rechazado')
    if rechazado == 'true':
        registros = registros.filter(rechazado=True)
    
    # Filtro por Punto de Internación (PI)
    internacion = request.GET.get('internacion')
    if internacion == 'true':
        registros = registros.filter(internacion=True)
    
    # Paginación
    paginator = Paginator(registros, 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Contar notificaciones no leídas para el usuario actual
    notificaciones_no_leidas = Notificacion.objects.filter(
        usuario=request.user,
        leida=False
    ).count()

    # Contar casos especiales pendientes (incluye urgentes sin_documento)
    casos_pendientes = CasoEspecial.objects.filter(estado='pendiente').count()
    casos_urgentes = CasoEspecial.objects.filter(estado='pendiente', razon='sin_documento').count()

    # Detectar si estamos en producción (cuando DJANGO_ENV != 'local')
    is_production = config('DJANGO_ENV', default='local') != 'local'

    context = {
        'page_obj': page_obj,
        'batches': batches,
        'is_superuser': request.user.is_superuser,
        'notificaciones_no_leidas': notificaciones_no_leidas,
        'casos_pendientes': casos_pendientes,
        'casos_urgentes': casos_urgentes,
        'is_production': is_production,
        'paises': get_paises(),
        'tiempos_rubros': [
            {'key': 'fma', 'label': 'FMA', 'personas': True},
            {'key': 'mexicanos', 'label': 'Mexicanos'},
            {'key': 'extranjeros', 'label': 'Extranjeros'},
        ],
    }

    return render(request, 'uploader/admin_list.html', context)


@flujo_principal_required
def generar_pin(request, fecha):
    """Vista para generar el PIN oficial del INM por fecha"""
    logger = logging.getLogger(__name__)
    
    # Convertir string de fecha a objeto date
    try:
        fecha_obj = datetime.strptime(fecha, '%Y-%m-%d').date()
    except ValueError as e:
        logger.error(f'Fecha inválida recibida: {fecha} - Error: {str(e)}')
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'error': 'Fecha inválida'}, status=400)
        messages.error(request, ' Fecha inválida.')
        return redirect('date_range_report')
    
    # Filtrar por rango de datetime en UTC para coincidir con cómo
    # date_range_report agrupa los registros (usa .date() sobre datetime UTC).
    # Evita el lookup __date que con USE_TZ=True convierte a la zona local
    # y desplaza la fecha de los registros antiguos (naive guardados como UTC).
    inicio_utc = datetime.combine(fecha_obj, time.min, tzinfo=dt_timezone.utc)
    fin_utc = inicio_utc + timedelta(days=1)
    registros_del_dia = Registro.objects.filter(
        vuelo_fecha__gte=inicio_utc,
        vuelo_fecha__lt=fin_utc,
    )

    if not registros_del_dia.exists():
        logger.warning(f'No se encontraron registros para la fecha: {fecha_obj}')
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'error': f'No se encontraron registros para la fecha {fecha_obj.strftime("%d/%m/%Y")}'}, status=404)
        messages.error(request, f' No se encontraron registros para la fecha {fecha_obj.strftime("%d/%m/%Y")}')
        return redirect('date_range_report')

    # Desglose por destino para el PIN Binacional.
    # Se combinan varias señales porque batch.tipo_vuelo puede estar en
    # NULL (lotes cargados antes de la migración 0008 o cuando el primer
    # registro no contenía TIJ/MEX): tipo_vuelo del batch, nombre del
    # archivo cargado y aeropuerto_llegada en sus variantes habituales.
    filtro_tij = (
        models.Q(batch__tipo_vuelo='PEK-TIJ') |
        models.Q(batch__archivo__icontains='PEK-TIJ') |
        models.Q(aeropuerto_llegada__icontains='TIJ') |
        models.Q(aeropuerto_llegada__icontains='TIJUANA') |
        models.Q(aeropuerto_llegada__icontains='蒂华纳')
    )
    filtro_mex = (
        models.Q(batch__tipo_vuelo='PEK-MEX') |
        models.Q(batch__archivo__icontains='PEK-MEX') |
        models.Q(aeropuerto_llegada__icontains='MEX') |
        models.Q(aeropuerto_llegada__icontains='MEXICO') |
        models.Q(aeropuerto_llegada__icontains='MÉXICO') |
        models.Q(aeropuerto_llegada__icontains='墨西哥')
    )

    # El PIN reporta sólo arribos a TIJ/MEX. Si el día incluye cargas de
    # regreso (vuelos de salida hacia el origen) sus pasajeros no deben
    # entrar a ningún conteo, porque inflan el total y no pertenecen al
    # desglose local/tránsito.
    registros_arribos = registros_del_dia.filter(filtro_tij | filtro_mex).distinct()

    total_pasajeros = registros_arribos.count()

    registros_sr = registros_arribos.filter(segunda_revision=True)
    total_sr = registros_sr.count()

    registros_internacion = registros_sr.filter(internacion=True)
    total_internaciones = registros_internacion.count()

    registros_rechazo = registros_sr.filter(rechazado=True)
    total_rechazos = registros_rechazo.count()

    # Conexiones = pasajeros PEK→MEX (tránsito) que no fueron rechazados
    registros_mex = registros_arribos.filter(filtro_mex)
    rechazados_mex = registros_mex.filter(rechazado=True).count()
    total_conexiones = registros_mex.count() - rechazados_mex

    total_pekin_tijuana = registros_arribos.filter(filtro_tij).count()
    total_pekin_mexico = registros_arribos.filter(filtro_mex).count()

    # Desglose por nacionalidad: mexicanos vs extranjeros, sobre ambos
    # Exceles del día (PEK-MEX y PEK-TIJ ya incluidos en registros_arribos).
    # Mexicano = documento con país de emisión México (código MEX/MX).
    # Cualquier otro código (chino, ruso, etc.) cuenta como extranjero.
    filtro_mexicano = (
        models.Q(codigo_pais_emision__iexact='MEX') |
        models.Q(codigo_pais_emision__iexact='MX')
    )
    total_mexicanos = registros_arribos.filter(filtro_mexicano).count()
    total_extranjeros = total_pasajeros - total_mexicanos

    # Vuelo y origen se autodetectan del primer arribo del día (lo que el
    # parser guardó al subir el Excel). Permite soportar rutas nuevas sin
    # tocar este código — basta con extender ORIGEN_POR_AEROPUERTO.
    vuelo_numero, origen_ciudad, origen_pais = _detectar_vuelo_y_origen(registros_arribos)
    
    # Datos completos de personas rechazadas
    rechazados_detalle = []
    for registro in registros_rechazo:
        # Obtener las URLs de las fotos de rechazo
        fotos_urls = []
        fotos_rechazo = registro.fotos_rechazo.all()
        for foto in fotos_rechazo:
            fotos_urls.append(request.build_absolute_uri(foto.foto.url))
        
        rechazados_detalle.append({
            'nombre': registro.nombre_pasajero,
            'genero': 'HOMBRE' if registro.genero == 'M' else 'MUJER' if registro.genero == 'F' else 'N/A',
            'nacionalidad': registro.pais_emision or 'N/A',
            'pasaporte': registro.numero_documento,
            'fecha_nacimiento': registro.fecha_nacimiento.strftime('%d.%m.%Y') if registro.fecha_nacimiento else 'N/A',
            'fotos': fotos_urls
        })
    
    # Si es una petición AJAX, devolver JSON
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        try:
            return JsonResponse({
                'fecha': fecha_obj.strftime('%Y-%m-%d'),
                'vuelo_numero': vuelo_numero,
                'origen_ciudad': origen_ciudad,
                'origen_pais': origen_pais,
                'total_pasajeros': total_pasajeros,
                'total_pekin_tijuana': total_pekin_tijuana,
                'total_pekin_mexico': total_pekin_mexico,
                'total_mexicanos': total_mexicanos,
                'total_extranjeros': total_extranjeros,
                'total_sr': total_sr,
                'total_internaciones': total_internaciones,
                'total_rechazos': total_rechazos,
                'total_conexiones': total_conexiones,
                'rechazados_detalle': rechazados_detalle,
            })
        except Exception as e:
            logger.error(f'Error al generar JSON del PIN: {str(e)}')
            return JsonResponse({'error': 'Error al generar el PIN'}, status=500)
    
    # Si no es AJAX, renderizar template completo (para compatibilidad)
    context = {
        'fecha': fecha_obj,
        'vuelo_numero': vuelo_numero,
        'origen_ciudad': origen_ciudad,
        'origen_pais': origen_pais,
        'total_pasajeros': total_pasajeros,
        'total_mexicanos': total_mexicanos,
        'total_extranjeros': total_extranjeros,
        'total_sr': total_sr,
        'total_internaciones': total_internaciones,
        'total_rechazos': total_rechazos,
        'total_conexiones': total_conexiones,
        'rechazados_detalle': rechazados_detalle,
        'is_superuser': request.user.is_superuser,
    }
    
    return render(request, 'uploader/pin_reporte.html', context)


DIAS_ES = {
    0: 'Lunes', 1: 'Martes', 2: 'Miércoles', 3: 'Jueves',
    4: 'Viernes', 5: 'Sábado', 6: 'Domingo'
}


def _compute_inadmitidos_data(fecha_inicio, fecha_fin):
    """Calcula los datos del reporte de inadmitidos para un rango de fechas."""
    filtro_tij = (
        models.Q(batch__tipo_vuelo='PEK-TIJ') |
        models.Q(batch__archivo__icontains='PEK-TIJ') |
        models.Q(aeropuerto_llegada__icontains='TIJ') |
        models.Q(aeropuerto_llegada__icontains='TIJUANA') |
        models.Q(aeropuerto_llegada__icontains='蒂华纳')
    )
    filtro_mex = (
        models.Q(batch__tipo_vuelo='PEK-MEX') |
        models.Q(batch__archivo__icontains='PEK-MEX') |
        models.Q(aeropuerto_llegada__icontains='MEX') |
        models.Q(aeropuerto_llegada__icontains='MEXICO') |
        models.Q(aeropuerto_llegada__icontains='MÉXICO') |
        models.Q(aeropuerto_llegada__icontains='墨西哥')
    )

    dias = []
    dia_actual = fecha_inicio
    while dia_actual <= fecha_fin:
        dias.append(dia_actual)
        dia_actual += timedelta(days=1)

    tiempos_map = {
        t.fecha: t for t in TiemposAtencion.objects.filter(fecha__gte=fecha_inicio, fecha__lte=fecha_fin)
    }

    dates, raw_dates = [], []
    totals_inadmitidos, totals_internaciones, totals_sr = [], [], []
    local, transito, total_pasajeros_list = [], [], []
    hora_inicio_list, hora_fin_list = [], []
    tiempo_extranjeros_list, tiempo_mexicanos_list, tiempo_fma_list = [], [], []
    fma_personas_list = []
    rs_hora_inicio_list, rs_hora_fin_list = [], []
    dur_fma_list, dur_mexicanos_list, dur_extranjeros_list = [], [], []
    dur_revisiones_secundarias_list = []
    nationalities_by_day = {}
    internaciones_by_day = {}
    motivos_rechazo = []

    def _min_de(t_obj):
        """Minutos desde medianoche de un `time`, o None si no hay valor."""
        return t_obj.hour * 60 + t_obj.minute if t_obj else None

    for i, dia in enumerate(dias):
        raw_dates.append(dia.strftime('%Y-%m-%d'))
        dates.append(f"{dia.day}-{MESES_ES[dia.month]}-{dia.strftime('%y')}")

        inicio_utc = datetime.combine(dia, time.min, tzinfo=dt_timezone.utc)
        fin_utc = inicio_utc + timedelta(days=1)
        registros_dia = Registro.objects.filter(
            vuelo_fecha__gte=inicio_utc,
            vuelo_fecha__lt=fin_utc,
        ).select_related('batch')
        # Sólo arribos: excluir vuelos de regreso (HU7926 MEX→PEK) que pueden
        # estar cargados el mismo día y que inflarían el total.
        registros_arribos = registros_dia.filter(filtro_tij | filtro_mex).distinct()

        inadmitidos = registros_arribos.filter(segunda_revision=True, rechazado=True)
        internados = registros_arribos.filter(segunda_revision=True, internacion=True)
        totals_inadmitidos.append(inadmitidos.count())
        totals_internaciones.append(internados.count())
        totals_sr.append(registros_arribos.filter(segunda_revision=True).count())
        local.append(registros_arribos.filter(filtro_tij).distinct().count())
        transito.append(registros_arribos.filter(filtro_mex).distinct().count())
        total_pasajeros_list.append(registros_arribos.count())

        t = tiempos_map.get(dia)
        hora_inicio_list.append(t.hora_inicio.strftime('%H:%M') if t else '')
        hora_fin_list.append(t.hora_fin.strftime('%H:%M') if t else '')

        def _dur_desde(inicio_min, fin_min):
            """Duración en minutos entre dos hitos (con cruce de medianoche)."""
            if inicio_min is None or fin_min is None:
                return ''
            d = fin_min - inicio_min
            if d < 0:
                d += 1440  # cruce de medianoche
            return d

        # FMA / Mexicanos / Extranjeros: hora de término; la duración SIEMPRE se
        # mide desde Hora Inicio (no en cascada entre rubros).
        inicio_min = _min_de(t.hora_inicio) if t else None
        rubros_dia = (
            ('fma', t.tiempo_fma if t else None),
            ('mexicanos', t.tiempo_mexicanos if t else None),
            ('extranjeros', t.tiempo_extranjeros if t else None),
        )
        horas_rubro, durs_rubro = {}, {}
        for nombre, obj in rubros_dia:
            horas_rubro[nombre] = obj.strftime('%H:%M') if obj else ''
            durs_rubro[nombre] = _dur_desde(inicio_min, _min_de(obj))

        # Revisiones Secundarias: ventana propia (inicio → fin).
        rs_ini = t.rs_hora_inicio if t else None
        rs_fin = t.rs_hora_fin if t else None
        dur_rs = _dur_desde(_min_de(rs_ini), _min_de(rs_fin))

        tiempo_fma_list.append(horas_rubro['fma'])
        tiempo_mexicanos_list.append(horas_rubro['mexicanos'])
        tiempo_extranjeros_list.append(horas_rubro['extranjeros'])
        # Conteo de personas FMA (independiente de las SR); None si no se capturó.
        fma_personas_list.append(t.fma_personas if (t and t.fma_personas is not None) else '')
        rs_hora_inicio_list.append(rs_ini.strftime('%H:%M') if rs_ini else '')
        rs_hora_fin_list.append(rs_fin.strftime('%H:%M') if rs_fin else '')
        dur_fma_list.append(durs_rubro['fma'])
        dur_mexicanos_list.append(durs_rubro['mexicanos'])
        dur_extranjeros_list.append(durs_rubro['extranjeros'])
        dur_revisiones_secundarias_list.append(dur_rs)

        nats_hoy = set()
        nat_counts = (
            inadmitidos
            .order_by()
            .values('pais_emision')
            .annotate(count=Count('id', distinct=True))
        )
        for item in nat_counts:
            nat = item['pais_emision'] or 'Sin especificar'
            nats_hoy.add(nat)
            if nat not in nationalities_by_day:
                nationalities_by_day[nat] = [0] * i
            nationalities_by_day[nat].append(item['count'])

        for nat in list(nationalities_by_day.keys()):
            if nat not in nats_hoy:
                nationalities_by_day[nat].append(0)

        # Desglose por nacionalidad de las internaciones (mismo patrón que arriba).
        nats_int_hoy = set()
        nat_int_counts = (
            internados
            .order_by()
            .values('pais_emision')
            .annotate(count=Count('id', distinct=True))
        )
        for item in nat_int_counts:
            nat = item['pais_emision'] or 'Sin especificar'
            nats_int_hoy.add(nat)
            if nat not in internaciones_by_day:
                internaciones_by_day[nat] = [0] * i
            internaciones_by_day[nat].append(item['count'])

        for nat in list(internaciones_by_day.keys()):
            if nat not in nats_int_hoy:
                internaciones_by_day[nat].append(0)

        if dia == fecha_fin:
            for registro in inadmitidos:
                comentario = (registro.comentario or '').strip()
                if comentario and comentario.lower() != 'none':
                    motivos_rechazo.append({
                        'nombre': registro.nombre_pasajero or 'Sin nombre',
                        'nacionalidad': registro.pais_emision or 'Sin especificar',
                        'numero_documento': registro.numero_documento or 'Sin documento',
                        'fecha': dates[i],
                        'comentario': comentario,
                    })

    # Autodetectar vuelo y origen del rango completo (primer arribo encontrado).
    inicio_rango_utc = datetime.combine(fecha_inicio, time.min, tzinfo=dt_timezone.utc)
    fin_rango_utc = datetime.combine(fecha_fin + timedelta(days=1), time.min, tzinfo=dt_timezone.utc)
    arribos_rango = Registro.objects.filter(
        vuelo_fecha__gte=inicio_rango_utc,
        vuelo_fecha__lt=fin_rango_utc,
    ).filter(filtro_tij | filtro_mex).distinct()
    vuelo_detectado, _ciudad_detectada, pais_detectado = _detectar_vuelo_y_origen(arribos_rango)

    return {
        'vuelo': vuelo_detectado,
        'origen': pais_detectado,
        'dates': dates,
        'raw_dates': raw_dates,
        'nationalities': nationalities_by_day,
        'totals_inadmitidos': totals_inadmitidos,
        'totals_internaciones': totals_internaciones,
        'internaciones_nationalities': internaciones_by_day,
        'totals_sr': totals_sr,
        'local': local,
        'transito': transito,
        'total_pasajeros': total_pasajeros_list,
        'hora_inicio': hora_inicio_list,
        'hora_fin': hora_fin_list,
        'tiempo_extranjeros': tiempo_extranjeros_list,
        'tiempo_mexicanos': tiempo_mexicanos_list,
        'tiempo_fma': tiempo_fma_list,
        'fma_personas': fma_personas_list,
        'rs_hora_inicio': rs_hora_inicio_list,
        'rs_hora_fin': rs_hora_fin_list,
        'dur_fma': dur_fma_list,
        'dur_mexicanos': dur_mexicanos_list,
        'dur_extranjeros': dur_extranjeros_list,
        'dur_revisiones_secundarias': dur_revisiones_secundarias_list,
        'motivos_rechazo': motivos_rechazo,
    }


@flujo_principal_required
def inadmitidos_page(request):
    """Página principal del reporte de inadmitidos"""
    return render(request, 'uploader/inadmitidos_report.html', {
        'is_superuser': request.user.is_superuser,
    })


@flujo_principal_required
def inadmitidos_data(request):
    """Endpoint AJAX que devuelve datos de inadmitidos por rango de fechas"""
    fecha_inicio_str = request.GET.get('fecha_inicio')
    fecha_fin_str = request.GET.get('fecha_fin')

    if not fecha_inicio_str or not fecha_fin_str:
        return JsonResponse({'error': 'Se requieren fecha_inicio y fecha_fin'}, status=400)

    try:
        fecha_inicio = datetime.strptime(fecha_inicio_str, '%Y-%m-%d').date()
        fecha_fin = datetime.strptime(fecha_fin_str, '%Y-%m-%d').date()
    except ValueError:
        return JsonResponse({'error': 'Formato de fecha inválido. Use YYYY-MM-DD'}, status=400)

    if fecha_fin < fecha_inicio:
        return JsonResponse({'error': 'La fecha fin debe ser mayor o igual a la fecha inicio'}, status=400)

    data = _compute_inadmitidos_data(fecha_inicio, fecha_fin)
    return JsonResponse(data)


@flujo_principal_required
def generar_inadmitidos_pdf(request):
    """Genera el reporte de inadmitidos como PDF con ReportLab"""
    from io import BytesIO
    from django.http import HttpResponse
    from reportlab.lib import colors
    from reportlab.lib.colors import HexColor
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

    fecha_inicio_str = request.GET.get('fecha_inicio')
    fecha_fin_str = request.GET.get('fecha_fin')

    if not fecha_inicio_str or not fecha_fin_str:
        return HttpResponse('Se requieren fecha_inicio y fecha_fin', status=400)

    try:
        fecha_inicio = datetime.strptime(fecha_inicio_str, '%Y-%m-%d').date()
        fecha_fin = datetime.strptime(fecha_fin_str, '%Y-%m-%d').date()
    except ValueError:
        return HttpResponse('Formato de fecha inválido', status=400)

    if fecha_fin < fecha_inicio:
        return HttpResponse('La fecha fin debe ser mayor o igual a la fecha inicio', status=400)

    data = _compute_inadmitidos_data(fecha_inicio, fecha_fin)
    n = len(data['dates'])
    nats = list(data['nationalities'].keys())

    # Colores
    ROJO = HexColor('#700606')
    ROSA = HexColor('#FED8D8')
    GRIS = HexColor('#DDDDDD')
    BLANCO_GRIS = HexColor('#F3F3F2')
    VERDE_INT = HexColor('#E2EFDA')
    VERDE_INT_TXT = HexColor('#375623')

    # Orientación según cantidad de días
    pagesize = landscape(letter) if n > 5 else letter
    page_w = pagesize[0] - 72  # ancho disponible con márgenes de 36pt

    label_w = 165
    date_col_w = max((page_w - label_w) / n, 55) if n > 0 else page_w
    col_widths = [label_w] + [date_col_w] * n

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=pagesize,
        leftMargin=36, rightMargin=36, topMargin=36, bottomMargin=36,
    )

    # ─── Filas de la tabla ───
    rows = []

    # Fila 0: encabezado Vuelo / Origen
    if n >= 3:
        header_row = ['Vuelo:', data['vuelo'], 'Origen:', data['origen']] + [''] * (n - 3)
    elif n == 2:
        header_row = ['Vuelo:', data['vuelo'], f"Origen: {data['origen']}"]
    else:
        header_row = [f"Vuelo: {data['vuelo']}   Origen: {data['origen']}"] + [''] * n
    rows.append(header_row)

    # Fila 1: INADMITIDOS (span completo)
    rows.append(['INADMITIDOS'] + [''] * n)

    # Fila 2: días de la semana
    day_names = [DIAS_ES[datetime.strptime(r, '%Y-%m-%d').date().weekday()] for r in data['raw_dates']]
    rows.append([''] + day_names)

    # Fila 3: Nacionalidad + fechas
    rows.append(['Nacionalidad'] + data['dates'])

    # Filas de nacionalidades
    NAT_START = 4
    for nat in nats:
        rows.append([nat] + [str(v) for v in data['nationalities'][nat]])

    nat_count = len(nats)

    # 2 filas vacías
    blank1 = len(rows); rows.append([''] * (n + 1))
    blank2 = len(rows); rows.append([''] * (n + 1))

    # Total Inadmitidos
    r_total_inad = len(rows)
    rows.append(['Total Inadmitidos'] + [str(v) for v in data['totals_inadmitidos']])

    # Separador
    r_sep1 = len(rows); rows.append([''] * (n + 1))

    # Total Internaciones
    r_total_int = len(rows)
    rows.append(['Total Internaciones:'] + [str(v) for v in data['totals_internaciones']])

    # Desglose por nacionalidad de las internaciones
    intern_nats = list(data.get('internaciones_nationalities', {}).keys())
    r_intern_start = len(rows)
    for nat in intern_nats:
        rows.append([nat] + [str(v) for v in data['internaciones_nationalities'][nat]])
    intern_count = len(intern_nats)

    # Total SR
    r_total_sr = len(rows)
    rows.append(['Total Segundas Revisiones:'] + [str(v) for v in data['totals_sr']])

    # Separador
    r_sep2 = len(rows); rows.append([''] * (n + 1))

    # Local / Tránsito / Total pasajeros
    r_local = len(rows); rows.append(['Local:'] + [str(v) for v in data['local']])
    r_trans = len(rows); rows.append(['En tr\xe1nsito:'] + [str(v) for v in data['transito']])
    r_total_pas = len(rows); rows.append(['Total pasajeros:'] + [str(v) for v in data['total_pasajeros']])

    # Separador antes de tiempos de atención
    r_sep3 = len(rows); rows.append([''] * (n + 1))

    # Helper para formatear la duración derivada (minutos → "Xh Ym").
    def _fmt_min(mins):
        try:
            n_ = int(mins)
        except (TypeError, ValueError):
            return '0m'
        if n_ <= 0:
            return '0m'
        h, m = divmod(n_, 60)
        if h == 0:
            return f'{m}m'
        if m == 0:
            return f'{h}h'
        return f'{h}h {m}m'

    def _val(x):
        s = str(x).strip()
        return int(s) if s.isdigit() else 0

    fma_vals = data.get('tiempo_fma', [''] * n)
    fma_personas = data.get('fma_personas', [''] * n)
    mex_vals = data.get('tiempo_mexicanos', [''] * n)
    ext_vals = data.get('tiempo_extranjeros', [''] * n)
    rs_ini_vals = data.get('rs_hora_inicio', [''] * n)
    rs_fin_vals = data.get('rs_hora_fin', [''] * n)
    dur_fma = data.get('dur_fma', [''] * n)
    dur_mex = data.get('dur_mexicanos', [''] * n)
    dur_ext = data.get('dur_extranjeros', [''] * n)
    dur_rs = data.get('dur_revisiones_secundarias', [''] * n)

    def _celda_rubro(hora, dur):
        """'HH:MM (Xh Ym)' con la duración derivada, o sólo la hora si no hay duración."""
        h = str(hora).strip()
        if not h:
            return ''
        return f'{h} ({_fmt_min(dur)})' if _val(dur) > 0 else h

    def _celda_dur(dur):
        """Sólo la duración formateada, o vacío si no hay."""
        return _fmt_min(dur) if _val(dur) > 0 else ''

    estilo_fma_cell = ParagraphStyle(
        'fma_cell', fontName='Helvetica', fontSize=7.5, leading=9, alignment=TA_CENTER
    )

    def _personas_txt(personas):
        """'<n> Personas' si hay conteo > 0, o vacío."""
        try:
            p = int(personas)
        except (TypeError, ValueError):
            return ''
        return f'{p} Personas' if p > 0 else ''

    def _celda_fma(hora, dur, personas):
        """Celda FMA: subdivide (personas / tiempo) cuando hay ambos; si sólo
        hay uno, muestra ese valor solo."""
        tiempo = _celda_rubro(hora, dur)
        pers = _personas_txt(personas)
        if pers and tiempo:
            return Paragraph(f'<b>{pers}</b><br/>{tiempo}', estilo_fma_cell)
        return pers or tiempo

    # Encabezado "Tiempos de atención" (span completo)
    r_tiempos_hdr = len(rows); rows.append(['Tiempos de atención'] + [''] * n)

    # Hora Inicio
    r_hora_ini = len(rows); rows.append(['Hora Inicio:'] + [str(v) for v in data.get('hora_inicio', [''] * n)])

    # FMA / Mexicanos / Extranjeros: hora de término + duración (desde Hora Inicio)
    # FMA además incluye el conteo de personas (celda subdividida si hay ambos).
    r_fma = len(rows); rows.append(['FMA:'] + [_celda_fma(fma_vals[i], dur_fma[i], fma_personas[i]) for i in range(n)])
    r_mex = len(rows); rows.append(['Mexicanos:'] + [_celda_rubro(mex_vals[i], dur_mex[i]) for i in range(n)])
    r_ext = len(rows); rows.append(['Extranjeros:'] + [_celda_rubro(ext_vals[i], dur_ext[i]) for i in range(n)])

    # Hora Fin (capturada aparte; se muestra después de los rubros)
    r_hora_fin = len(rows); rows.append(['Hora Fin:'] + [str(v) for v in data.get('hora_fin', [''] * n)])

    # Revisiones Secundarias: ventana propia (Inicio / Fin / Duración)
    r_rs_ini = len(rows); rows.append(['RS Hora Inicio:'] + [str(rs_ini_vals[i]) for i in range(n)])
    r_rs_fin = len(rows); rows.append(['RS Hora Fin:'] + [str(rs_fin_vals[i]) for i in range(n)])
    r_rs_dur = len(rows); rows.append(['RS Duración:'] + [_celda_dur(dur_rs[i]) for i in range(n)])

    # ─── Estilos ───
    cmd = [
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('LEFTPADDING', (0, 0), (-1, -1), 5),
        ('RIGHTPADDING', (0, 0), (-1, -1), 5),

        # Fila 0: encabezado rojo
        ('BACKGROUND', (0, 0), (-1, 0), ROJO),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),

        # Fila 1: INADMITIDOS rosado
        ('SPAN', (0, 1), (-1, 1)),
        ('BACKGROUND', (0, 1), (-1, 1), ROSA),
        ('TEXTCOLOR', (0, 1), (-1, 1), ROJO),
        ('FONTNAME', (0, 1), (-1, 1), 'Helvetica-Bold'),
        ('ALIGN', (0, 1), (-1, 1), 'CENTER'),

        # Fila 2: días de semana rojo
        ('BACKGROUND', (0, 2), (-1, 2), ROJO),
        ('TEXTCOLOR', (0, 2), (-1, 2), colors.white),
        ('FONTNAME', (0, 2), (-1, 2), 'Helvetica-Bold'),
        ('ALIGN', (0, 2), (-1, 2), 'CENTER'),

        # Fila 3: Nacionalidad/fechas rosado
        ('BACKGROUND', (0, 3), (-1, 3), ROSA),
        ('TEXTCOLOR', (0, 3), (-1, 3), ROJO),
        ('FONTNAME', (0, 3), (-1, 3), 'Helvetica-Bold'),
        ('ALIGN', (0, 3), (-1, 3), 'CENTER'),
        ('ALIGN', (0, 3), (0, 3), 'LEFT'),

        # Total Inadmitidos rojo
        ('BACKGROUND', (0, r_total_inad), (-1, r_total_inad), ROJO),
        ('TEXTCOLOR', (0, r_total_inad), (-1, r_total_inad), colors.white),
        ('FONTNAME', (0, r_total_inad), (-1, r_total_inad), 'Helvetica-Bold'),

        # Total Internaciones / SR rosado
        ('BACKGROUND', (0, r_total_int), (-1, r_total_int), ROSA),
        ('TEXTCOLOR', (0, r_total_int), (-1, r_total_int), ROJO),
        ('FONTNAME', (0, r_total_int), (-1, r_total_int), 'Helvetica-Bold'),
        ('BACKGROUND', (0, r_total_sr), (-1, r_total_sr), ROSA),
        ('TEXTCOLOR', (0, r_total_sr), (-1, r_total_sr), ROJO),
        ('FONTNAME', (0, r_total_sr), (-1, r_total_sr), 'Helvetica-Bold'),

        # Local / Tránsito blanco-gris
        ('BACKGROUND', (0, r_local), (-1, r_local), BLANCO_GRIS),
        ('BACKGROUND', (0, r_trans), (-1, r_trans), BLANCO_GRIS),

        # Total pasajeros gris + negrita
        ('BACKGROUND', (0, r_total_pas), (-1, r_total_pas), GRIS),
        ('FONTNAME', (0, r_total_pas), (-1, r_total_pas), 'Helvetica-Bold'),

        # Separadores y filas vacías en blanco-gris
        ('BACKGROUND', (0, blank1), (-1, blank1), BLANCO_GRIS),
        ('BACKGROUND', (0, blank2), (-1, blank2), BLANCO_GRIS),
        ('BACKGROUND', (0, r_sep1), (-1, r_sep1), BLANCO_GRIS),
        ('BACKGROUND', (0, r_sep2), (-1, r_sep2), BLANCO_GRIS),
        ('BACKGROUND', (0, r_sep3), (-1, r_sep3), BLANCO_GRIS),

        # Hora Inicio / Hora Fin
        ('BACKGROUND', (0, r_hora_ini), (-1, r_hora_ini), BLANCO_GRIS),
        ('BACKGROUND', (0, r_hora_fin), (-1, r_hora_fin), BLANCO_GRIS),
        ('FONTNAME', (0, r_hora_ini), (0, r_hora_ini), 'Helvetica-Bold'),
        ('FONTNAME', (0, r_hora_fin), (0, r_hora_fin), 'Helvetica-Bold'),

        # Encabezado "Tiempos de atención" rojo span
        ('SPAN', (0, r_tiempos_hdr), (-1, r_tiempos_hdr)),
        ('BACKGROUND', (0, r_tiempos_hdr), (-1, r_tiempos_hdr), ROJO),
        ('TEXTCOLOR', (0, r_tiempos_hdr), (-1, r_tiempos_hdr), colors.white),
        ('FONTNAME', (0, r_tiempos_hdr), (-1, r_tiempos_hdr), 'Helvetica-Bold'),
        ('ALIGN', (0, r_tiempos_hdr), (-1, r_tiempos_hdr), 'CENTER'),

        # FMA / Mexicanos / Extranjeros
        ('BACKGROUND', (0, r_fma), (-1, r_fma), BLANCO_GRIS),
        ('BACKGROUND', (0, r_mex), (-1, r_mex), GRIS),
        ('BACKGROUND', (0, r_ext), (-1, r_ext), BLANCO_GRIS),
        ('FONTNAME', (0, r_fma), (0, r_fma), 'Helvetica-Bold'),
        ('FONTNAME', (0, r_mex), (0, r_mex), 'Helvetica-Bold'),
        ('FONTNAME', (0, r_ext), (0, r_ext), 'Helvetica-Bold'),

        # Revisiones Secundarias: Inicio / Fin / Duración
        ('BACKGROUND', (0, r_rs_ini), (-1, r_rs_ini), BLANCO_GRIS),
        ('BACKGROUND', (0, r_rs_fin), (-1, r_rs_fin), BLANCO_GRIS),
        ('BACKGROUND', (0, r_rs_dur), (-1, r_rs_dur), GRIS),
        ('FONTNAME', (0, r_rs_ini), (0, r_rs_ini), 'Helvetica-Bold'),
        ('FONTNAME', (0, r_rs_fin), (0, r_rs_fin), 'Helvetica-Bold'),
        ('FONTNAME', (0, r_rs_dur), (-1, r_rs_dur), 'Helvetica-Bold'),
    ]

    # Span del origen en el encabezado si hay columnas suficientes
    if n >= 3:
        cmd.append(('SPAN', (3, 0), (-1, 0)))

    # Colores alternos en filas de nacionalidades
    for i in range(nat_count):
        row_idx = NAT_START + i
        bg = BLANCO_GRIS if i % 2 == 0 else GRIS
        cmd.append(('BACKGROUND', (0, row_idx), (-1, row_idx), bg))
        cmd.append(('FONTNAME', (0, row_idx), (0, row_idx), 'Helvetica-Bold'))

    # Desglose de internaciones por nacionalidad (verde claro, etiqueta indentada)
    for i in range(intern_count):
        row_idx = r_intern_start + i
        cmd.append(('BACKGROUND', (0, row_idx), (-1, row_idx), VERDE_INT))
        cmd.append(('TEXTCOLOR', (0, row_idx), (-1, row_idx), VERDE_INT_TXT))
        cmd.append(('FONTNAME', (0, row_idx), (0, row_idx), 'Helvetica-Bold'))
        cmd.append(('LEFTPADDING', (0, row_idx), (0, row_idx), 18))

    tabla = Table(rows, colWidths=col_widths)
    tabla.setStyle(TableStyle(cmd))

    elements = [tabla]

    # ─── Motivos de rechazo ───
    if data['motivos_rechazo']:
        elements.append(Spacer(1, 10))
        estilo_normal = ParagraphStyle('mn', fontName='Helvetica', fontSize=8, leading=12)
        motivos_rows = []
        for m in data['motivos_rechazo']:
            texto = f"<b>Extranjero de {m['nacionalidad']}</b> [{m['fecha']}] {m['nombre']}, {m['numero_documento']}: {m['comentario']}"
            motivos_rows.append([Paragraph(texto, estilo_normal)])

        encabezado_style = ParagraphStyle('mh', fontName='Helvetica-Bold', fontSize=8, leading=12)
        motivos_rows.insert(0, [Paragraph('Motivos de rechazo del día:', encabezado_style)])

        t_motivos = Table(motivos_rows, colWidths=[page_w])
        t_motivos.setStyle(TableStyle([
            ('BOX', (0, 0), (-1, -1), 1, colors.black),
            ('BACKGROUND', (0, 0), (-1, -1), BLANCO_GRIS),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('LEFTPADDING', (0, 0), (-1, -1), 8),
            ('RIGHTPADDING', (0, 0), (-1, -1), 8),
        ]))
        elements.append(t_motivos)

    doc.build(elements)
    pdf_bytes = buffer.getvalue()
    buffer.close()

    nombre = f"Inadmitidos_{data['vuelo']}_{fecha_inicio_str}_{fecha_fin_str}.pdf"
    response = HttpResponse(pdf_bytes, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{nombre}"'
    return response


@flujo_principal_required
def generar_inadmitidos_excel(request):
    """Genera el reporte de inadmitidos como Excel (.xlsx) con openpyxl.

    Replica el layout de la tabla del reporte; la cantidad de columnas de
    fechas depende del rango consultado (un día por columna)."""
    from io import BytesIO
    from urllib.parse import quote
    from django.http import HttpResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    fecha_inicio_str = request.GET.get('fecha_inicio')
    fecha_fin_str = request.GET.get('fecha_fin')

    if not fecha_inicio_str or not fecha_fin_str:
        return HttpResponse('Se requieren fecha_inicio y fecha_fin', status=400)

    try:
        fecha_inicio = datetime.strptime(fecha_inicio_str, '%Y-%m-%d').date()
        fecha_fin = datetime.strptime(fecha_fin_str, '%Y-%m-%d').date()
    except ValueError:
        return HttpResponse('Formato de fecha inválido', status=400)

    if fecha_fin < fecha_inicio:
        return HttpResponse('La fecha fin debe ser mayor o igual a la fecha inicio', status=400)

    # El Excel "Autoridades" omite la sección de Tiempos de Atención y los motivos.
    autoridades = request.GET.get('autoridades') == '1'

    data = _compute_inadmitidos_data(fecha_inicio, fecha_fin)
    n = len(data['dates'])
    nats = list(data['nationalities'].keys())

    # La columna Total sólo aporta con varios días (con uno solo sería redundante).
    incluir_total = n > 1
    last_col = (n + 2) if incluir_total else (n + 1)  # etiqueta + n días [+ Total]

    # ─── Colores tomados tal cual del CSS del reporte (inadmitidos_report.css),
    #     aplanando los rgba() sobre fondo blanco. Prefijo 'FF' = alfa opaco. ───
    MAROON       = 'FF6B1D1D'  # row-vuelo / row-dias
    DARK_RED     = 'FF4D060A'  # total-sr / encabezado de sección
    ROSA_TITULO  = 'FFFFEEED'  # row-title / row-total-inad
    ROSA_FECHAS  = 'FFFFF7F6'  # row-fechas
    VERDE_INTERN = 'FFE6EDE9'  # row-total-intern (el verde va aquí, en el Total)
    DATA_BG      = 'FFF9F9FC'  # row-data / row-intern-nac
    PAX_BG       = 'FFE2E2E5'  # row-total-pax

    fill_maroon   = PatternFill('solid', fgColor=MAROON)
    fill_dark_red = PatternFill('solid', fgColor=DARK_RED)
    fill_rosa_tit = PatternFill('solid', fgColor=ROSA_TITULO)
    fill_rosa_fec = PatternFill('solid', fgColor=ROSA_FECHAS)
    fill_verde    = PatternFill('solid', fgColor=VERDE_INTERN)
    fill_data     = PatternFill('solid', fgColor=DATA_BG)
    fill_pax      = PatternFill('solid', fgColor=PAX_BG)

    # Tipografía Calibri en todo (igual que el reporte)
    font_data   = Font(name='Calibri', size=11, color='FF1A1C1E')              # valores / etiquetas normales
    font_blanco = Font(name='Calibri', size=11, bold=True, color='FFFFFFFF')   # filas maroon / encabezados
    font_maroon = Font(name='Calibri', size=11, bold=True, color='FF6B1D1D')   # row-title / motivos
    font_fechas = Font(name='Calibri', size=10, bold=True, color='FF554241')   # fila de fechas
    font_pax    = Font(name='Calibri', size=11, bold=True, color='FF1A1C1E')   # total pasajeros

    # Filas de totales destacados: Calibri 12 en negritas
    font_tot_inad   = Font(name='Calibri', size=12, bold=True, color='FF6B1D1D')   # Total Inadmitidos
    font_tot_intern = Font(name='Calibri', size=12, bold=True, color='FF1C2522')   # Total Internaciones
    font_tot_sr     = Font(name='Calibri', size=12, bold=True, color='FFFFFFFF')   # Total Segundas Revisiones

    thin = Side(style='thin', color='FFE2E2E5')
    borde = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left = Alignment(horizontal='left', vertical='center', wrap_text=True)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Inadmitidos'

    def _fmt_min(mins):
        try:
            m_ = int(mins)
        except (TypeError, ValueError):
            return '0m'
        if m_ <= 0:
            return '0m'
        h, m = divmod(m_, 60)
        if h == 0:
            return f'{m}m'
        if m == 0:
            return f'{h}h'
        return f'{h}h {m}m'

    def _val(x):
        s = str(x).strip()
        return int(s) if s.isdigit() else 0

    def _num(v):
        """Devuelve int cuando el valor es un entero, para que Excel lo trate
        como número (sin el aviso 'número almacenado como texto'); si no, el
        texto original (p. ej. horas '17:36' o duraciones '24m')."""
        if isinstance(v, bool):
            return v
        if isinstance(v, int):
            return v
        s = str(v).strip()
        return int(s) if s.lstrip('-').isdigit() else v

    def _celda_rubro(hora, dur):
        h = str(hora).strip()
        if not h:
            return ''
        return f'{h} ({_fmt_min(dur)})' if _val(dur) > 0 else h

    def _celda_dur(dur):
        return _fmt_min(dur) if _val(dur) > 0 else ''

    def _celda_fma(hora, dur, personas):
        tiempo = _celda_rubro(hora, dur)
        try:
            p = int(personas)
        except (TypeError, ValueError):
            p = 0
        pers = f'{p} Personas' if p > 0 else ''
        if pers and tiempo:
            return f'{pers}\n{tiempo}'
        return pers or tiempo

    def _suma(arr):
        """Suma horizontal de conteos (la columna Total)."""
        return sum(_val(v) for v in arr)

    def _suma_dur(arr):
        """Suma de duraciones (minutos) formateada, o vacío si no hay."""
        tot = 0
        for v in arr:
            try:
                iv = int(v)
            except (TypeError, ValueError):
                iv = 0
            if iv > 0:
                tot += iv
        return _fmt_min(tot) if tot > 0 else ''

    def _con_total(valores, total):
        """Agrega la celda Total al final de la fila (sólo si incluir_total)."""
        return valores + [total] if incluir_total else valores

    # Cada elemento de `filas` es (lista_de_valores, fill, font, span_completo).
    # span_completo=True fusiona toda la fila en una sola celda.
    filas = []

    # Encabezado Vuelo / Origen (row-vuelo, maroon). Con ancho suficiente se
    # ponen en celdas separadas (y se fusiona el origen); si no, en una sola.
    if last_col >= 4:
        filas.append((
            ['Vuelo:', data['vuelo'], 'Origen:', data['origen']] + [''] * (last_col - 4),
            fill_maroon, font_blanco, False,
        ))
        merge_origen = True
    else:
        filas.append(([f"Vuelo: {data['vuelo']}   Origen: {data['origen']}"], fill_maroon, font_blanco, True))
        merge_origen = False

    # INADMITIDOS (row-title, rosa claro, span)
    filas.append((['INADMITIDOS'], fill_rosa_tit, font_maroon, True))
    # Días de la semana (row-dias, maroon) + encabezado de la columna Total
    day_names = [DIAS_ES[datetime.strptime(r, '%Y-%m-%d').date().weekday()] for r in data['raw_dates']]
    filas.append((_con_total(['Nacionalidad'] + day_names, 'Total'), fill_maroon, font_blanco, False))
    # Fechas (row-fechas, rosa muy claro)
    filas.append((_con_total([''] + list(data['dates']), ''), fill_rosa_fec, font_fechas, False))

    # Nacionalidades de inadmitidos: SÓLO la celda del nombre va en rosa (5º
    # elemento = relleno de la etiqueta); los números quedan neutros. Cada fila
    # se enmarca después (box_rows) para focalizarla.
    box_rows = []
    for nat in nats:
        vals = data['nationalities'][nat]
        filas.append((_con_total([nat] + [_num(v) for v in vals], _suma(vals)), fill_data, font_data, False, fill_rosa_tit))
        box_rows.append(len(filas))  # nº de fila en la hoja (el volcado usa start=1)

    # Total Inadmitidos (banda rosa completa, Calibri 12 negritas, enmarcada)
    filas.append((_con_total(['Total Inadmitidos'] + [_num(v) for v in data['totals_inadmitidos']],
                             _suma(data['totals_inadmitidos'])), fill_rosa_tit, font_tot_inad, False))
    box_rows.append(len(filas))
    # Total Internaciones (banda verde completa, Calibri 12 negritas, enmarcada)
    filas.append((_con_total(['Total Internaciones:'] + [_num(v) for v in data['totals_internaciones']],
                             _suma(data['totals_internaciones'])), fill_verde, font_tot_intern, False))
    box_rows.append(len(filas))
    # Desglose internaciones: SÓLO la celda del nombre va en verde; enmarcada.
    intern_nats = list(data.get('internaciones_nationalities', {}).keys())
    for nat in intern_nats:
        vals = data['internaciones_nationalities'][nat]
        filas.append((_con_total([f'   {nat}'] + [_num(v) for v in vals], _suma(vals)), fill_data, font_data, False, fill_verde))
        box_rows.append(len(filas))
    # Total Segundas Revisiones (row-total-sr, rojo oscuro, Calibri 12 negritas)
    filas.append((_con_total(['Total Segundas Revisiones:'] + [_num(v) for v in data['totals_sr']],
                             _suma(data['totals_sr'])), fill_dark_red, font_tot_sr, False))
    # Local / Tránsito (row-data) / Total pasajeros (row-total-pax)
    filas.append((_con_total(['Local:'] + [_num(v) for v in data['local']], _suma(data['local'])), fill_data, font_data, False))
    filas.append((_con_total(['En tránsito:'] + [_num(v) for v in data['transito']], _suma(data['transito'])), fill_data, font_data, False))
    filas.append((_con_total(['Total pasajeros:'] + [_num(v) for v in data['total_pasajeros']],
                             _suma(data['total_pasajeros'])), fill_pax, font_pax, False))

    # ─── Tiempos de atención (sólo en el Excel completo, no en el de autoridades) ───
    if not autoridades:
        fma_vals = data.get('tiempo_fma', [''] * n)
        fma_personas = data.get('fma_personas', [''] * n)
        mex_vals = data.get('tiempo_mexicanos', [''] * n)
        ext_vals = data.get('tiempo_extranjeros', [''] * n)
        rs_ini_vals = data.get('rs_hora_inicio', [''] * n)
        rs_fin_vals = data.get('rs_hora_fin', [''] * n)
        dur_fma = data.get('dur_fma', [''] * n)
        dur_mex = data.get('dur_mexicanos', [''] * n)
        dur_ext = data.get('dur_extranjeros', [''] * n)
        dur_rs = data.get('dur_revisiones_secundarias', [''] * n)

        filas.append((['Tiempos de atención'], fill_dark_red, font_blanco, True))
        filas.append((_con_total(['Hora Inicio:'] + [str(v) for v in data.get('hora_inicio', [''] * n)], ''), fill_data, font_data, False))
        filas.append((_con_total(['FMA:'] + [_celda_fma(fma_vals[i], dur_fma[i], fma_personas[i]) for i in range(n)], _suma_dur(dur_fma)), fill_data, font_data, False))
        filas.append((_con_total(['Mexicanos:'] + [_celda_rubro(mex_vals[i], dur_mex[i]) for i in range(n)], _suma_dur(dur_mex)), fill_data, font_data, False))
        filas.append((_con_total(['Extranjeros:'] + [_celda_rubro(ext_vals[i], dur_ext[i]) for i in range(n)], _suma_dur(dur_ext)), fill_data, font_data, False))
        filas.append((_con_total(['Hora Fin:'] + [str(v) for v in data.get('hora_fin', [''] * n)], ''), fill_data, font_data, False))
        filas.append((_con_total(['RS Hora Inicio:'] + [str(rs_ini_vals[i]) for i in range(n)], ''), fill_data, font_data, False))
        filas.append((_con_total(['RS Hora Fin:'] + [str(rs_fin_vals[i]) for i in range(n)], ''), fill_data, font_data, False))
        filas.append((_con_total(['RS Duración:'] + [_celda_dur(dur_rs[i]) for i in range(n)], _suma_dur(dur_rs)), fill_data, font_data, False))

    # ─── Volcado a la hoja ───
    # Cada fila es (valores, fill, font, span[, label_fill]); si label_fill está
    # presente, sólo la celda de la etiqueta (col A) usa ese color.
    for r_idx, fila in enumerate(filas, start=1):
        valores, fill, font, span = fila[0], fila[1], fila[2], fila[3]
        label_fill = fila[4] if len(fila) > 4 else None
        for c_idx in range(1, last_col + 1):
            valor = valores[c_idx - 1] if c_idx - 1 < len(valores) else ''
            celda = ws.cell(row=r_idx, column=c_idx, value=valor)
            celda.fill = label_fill if (label_fill is not None and c_idx == 1) else fill
            celda.font = font
            celda.border = borde
            celda.alignment = left if c_idx == 1 else center
        if span:
            ws.merge_cells(start_row=r_idx, start_column=1, end_row=r_idx, end_column=last_col)
            ws.cell(row=r_idx, column=1).alignment = center

    # Span del origen en el encabezado (col D..fin) cuando van en celdas separadas
    if merge_origen:
        ws.merge_cells(start_row=1, start_column=4, end_row=1, end_column=last_col)
        ws.cell(row=1, column=4).alignment = left

    # Anchos de columna
    ws.column_dimensions['A'].width = 26
    for c in range(2, last_col + 1):
        ws.column_dimensions[get_column_letter(c)].width = 16

    # ── Enmarca cada fila de nacionalidades y de totales (Inadmitidos /
    #    Internaciones) para focalizarlas: borde superior/inferior en toda la
    #    fila, izquierdo en la etiqueta y derecho hasta la última columna
    #    (Total). Verticales internos quedan tenues. ──
    box = Side(style='thin', color='FF000000')
    for row_idx in box_rows:
        for c in range(1, last_col + 1):
            izq = box if c == 1 else thin
            der = box if c == last_col else thin
            ws.cell(row=row_idx, column=c).border = Border(left=izq, right=der, top=box, bottom=box)

    # ─── Motivos de rechazo (no se incluyen en el Excel de autoridades) ───
    if data['motivos_rechazo'] and not autoridades:
        r = len(filas) + 2
        c = ws.cell(row=r, column=1, value='Motivos de rechazo del día:')
        c.font = font_maroon
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=last_col)
        for m in data['motivos_rechazo']:
            r += 1
            texto = (f"Extranjero de {m['nacionalidad']} [{m['fecha']}] "
                     f"{m['nombre']}, {m['numero_documento']}: {m['comentario']}")
            cm = ws.cell(row=r, column=1, value=texto)
            cm.font = font_data
            cm.alignment = left
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=last_col)

    buffer = BytesIO()
    wb.save(buffer)
    xlsx_bytes = buffer.getvalue()
    buffer.close()

    prefijo = 'Inadmitidos_Autoridades' if autoridades else 'Inadmitidos'
    nombre = f"{prefijo}_{data['vuelo']}_{fecha_inicio_str}_{fecha_fin_str}.xlsx"
    response = HttpResponse(
        xlsx_bytes,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f"attachment; filename*=UTF-8''{quote(nombre)}"
    return response